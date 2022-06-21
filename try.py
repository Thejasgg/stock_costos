from ast import Break
import datetime
from datetime import date, time, datetime
import pandas as pd
from pandas import ExcelWriter
from seaborn import load_dataset
import openpyxl as pn
from openpyxl import Workbook
from openpyxl import load_workbook

archiveStock = "Base de datos de stock.xlsx"
archive = pd.read_excel(archiveStock)
#archiveSales = "Ventas.xlsx"
#archivesales = pd.read_excel(archiveSales)
#def calculatorMissing(inventary):
#    inventary = inventary - 1
#    return inventary
Dictionary1 = { 'A': 'Geeks', 'B': 'For'}
 
Third_value = Dictionary1.setdefault('C')
print("Dictionary:", Dictionary1)
print("Third_value:", Third_value)
 
Fourth_value = Dictionary1.setdefault('D', 'Geeks')
print("Dictionary:", Dictionary1)
print("Fourth_value:", Fourth_value)
#archive["inventary"] = archive["inventary"].apply(calculatorMissing)
#archive.loc[archive["code"]==30, "inventary"]+= 1
archiveForSales = "Ventas.xlsx"
archiveSales = pd.read_excel(archiveForSales)
#Diccionario de datos de las ventas
sales = {"Code": "","Product": "","Cant": "", "Cost": "","Price": "", "Ingres": "","Total Cost": "", "Benefits": ""}
#Diccionario del resumen diario de ventas
resume = {"Date": "", "Tipe": "", "Ingres": "", "Cost": "", "Benefit": ""}
#Introduce el código para buscar dentro del stock
introduceCode = int(input("Escribe el código de un producto: "))
#Condición para escribir el excel
code = archive.loc[archive["code"]==introduceCode]
#Escribe las ventas
sales["Code"] = introduceCode
sales["Product"] = code["product"]
sales["Cant"] = int(input("Escribe la cantidad del producto: "))
sales["Cost"] = code["cost"]
sales["Price"] = code["price"]
sales["Ingres"] = sales["Cant"] * sales["Price"]
sales["Total Cost"] = sales["Cant"] * sales["Cost"]
sales["Benefits"] = sales["Ingres"] - (sales["Cost"] * sales["Cant"])
#Actualiza el stock
archive.loc[archive["code"]==introduceCode, "inventary"]-= sales["Cant"]
dataSales = pd.DataFrame(sales)
dataSales.to_excel("Resumen.xlsx", sheet_name="Resumen", index=False)

#Agregar al main para guardar el cambio en el stock
archiveForSales = "Ventas.xlsx"
archiveSales = pd.read_excel(archiveForSales)
archiveSales = archiveSales.append(dataSales, ignore_index=True)
print(archiveSales)
archiveSales.to_excel("Ventas.xlsx")
#save()
#crea el dataframe de las ventas
#dataSales = pd.DataFrame(sales)
#busca la fecha actual
#dateToday = datetime.now()
#print(dateToday)
#Aplica condiciones para la fecha actual
#dateToday = dateToday.strftime("%d/%m/%Y")
#print(dateToday)
#Rellena el diccionario resumen con datos generados en el stock y las ventas
#resume["Date"]= dateToday
#resume["Tipe"]= "Ventas"
#resume["Ingres"] = sales["Ingres"]
#resume["Cost"] = sales["Total Cost"]
#resume["Benefit"] = resume["Ingres"] - resume["Cost"]
#Convierte el diccionario en dataframe
#dataResume = pd.DataFrame(resume)
#print(dataResume)
#crea las ventas en excel
#Crea el resumen de ventas en excel
#dataResume.to_excel("Resumen.xlsx", sheet_name="Resumen", index=False)


        #abrir el libro de trabajo
        #orkBook0 = Workbook()
        #guardar el libro
        #workBook0.save("Ventas.xlsx")
        #cargar el libro
        #workBook = load_workbook("Ventas.xlsx")
        #workSheet1 = workBook.create_sheet("Ventas")
        #cambiar el nombre al archivo
        #workSheet = workBook["Sheet"]
        #workSheet.title = "Ventas"
        #agregar una hoja
        #workSheet2 = workBook.create_sheet("Resumen")
        #guardar el archivo
        #workBook.save("Ventas.xlsx")
        




        #dataSales = pd.DataFrame(sales)
        #print(optionsExcelSales)
        #"""optionsExcelSalesUser = int(input("Elige una opción: "))
        #if optionsExcelSalesUser == 1:
            #createArchive()
        #elif optionsExcelSalesUser == 2:
            #archive_sales = "Ventas.xlsx"
            #archiveSales = pd.read_excel(archive_sales)
            #archiveSales = archiveSales.append(sales,ignore_index=True)
            #saveSales()
        #elif optionsExcelSalesUser == 3:
            #Break
        #else:
           #print("Elige correctamente")

#optionsExcelSales= """
#1)_ Crea el archivo y sustituye el anterior
#2)_ Guarda el archivo
#3)_ Cierra el programa
#"""
#"""

#writer = pd.ExcelWriter('Ventas.xlsx')
#dataSales.to_excel(writer, sheet_name="ventas", index=False)
#dataResume.to_excel(writer, sheet_name="resumen", index=False)
#writer.save()
#writer.close()
    
#writer = pd.ExcelWriter('Ventas.xlsx')
#dataSales.to_excel(writer, sheet_name="ventas", index=False)
#dataResume.to_excel(writer, sheet_name="resumen", index=False)
#writer.save()
#writer.close()
#sales.dict_excel("c:/Users/Usuario/Documents/Python/Proyecto final/ejemplo_python2.xlsl")

#archive_sales = archive_sales.append(sales, ignore_index=True)
#print(archive)
#print(code)


##d_index = archive.to_dict('index')
#print(d_index)
#Sirve para filtrar el contenido de una columna
#archive = archive["price"] 

#Fitrar varias columnas
#archive = archive[["code", "product", "price"]]

#Filtrar por filas (con solo un indice me muestra una sola fila, con 2 indices me muestra hasta la anterior
# del máximo)
#archive = archive.iloc[0:3]
#Filtrar por contenido por fila pero eligiendola
#archive = archive.iloc[[0,5,2]]

#Filtrar por contenido en la primera fila
#archive = archive.loc[[0,5,4]]

#Filtrar por filas y columnas
#archive = archive.loc[[0,5,4] , ["code", "product"]]

#Filtrar por condición de una columna
#archive = archive[archive["product"]=="Alfajor Fantoche"]

#Filtrar por el nombre del producto
#introduceCode = str(input("Escribe el nombre de un producto"))
#archive = archive[archive["product"]==str(introduceCode)]

#Filtrar por producto
#introduceCode = str(input("Escribe el código de un producto"))
#archive = archive[archive["product"]==str(introduceCode)]
#line = int(input("Escriba el número de la fila que desea filtrar"))
#archive = archive.iloc[int(line)]
# column = str(input("Escribe el nombre de la columna que deseas filtrar"))
# archive = archive[column]
# #def calculatorBenefit ():
#     #moneyBenefit = ("price" - "cost") * "cant sells"
#     #print(moneyBenefit)

# print(archive)