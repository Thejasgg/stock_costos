from ast import Break
import datetime
from datetime import date, time, datetime
import pandas as pd
from pandas import ExcelWriter
from seaborn import load_dataset
import openpyxl as pn
from openpyxl import Workbook
from openpyxl import load_workbook



#abrir el libro de trabajo
#workBook0 = Workbook()
#guardar el libro
#workBook0.save("Ventas.xlsx")
#cargar el libro
#workBook = load_workmbre al archivo
#workSheet = workBook["Sheet"]
#workSheet.title = "Ventas"
#agregar una hoja
#workSheet2 = workBook.create_sheet("Resumen")
#guardar el archivo
#workBook.save("Ventas.xlsx")
archiveStock = "Base de datos de stock.xlsx"
archive = pd.read_excel(archiveStock)
archiveForSales = "Ventas.xlsx"
archiveSales = pd.read_excel(archiveForSales)

introduceCode = int(input("Escribe el código de un producto: "))
sales = {"date" : "", "Code": "","Product": "","Cant": "", "Cost": "","Price": "", "Ingres": "","Total Cost": "", "Benefits": ""}
code = archive.loc[archive["code"]==introduceCode]
dateToday = datetime.now()
dateToday = dateToday.strftime("%d/%m/%Y")
sales["Date"]= dateToday
sales["Code"] = introduceCode
sales["Product"] = code["product"]
sales["Cant"] = int(input("Escribe la cantidad del producto: "))
sales["Cost"] = code["cost"]
sales["Price"] = code["price"]
sales["Ingres"] = sales["Cant"] * sales["Price"]
sales["Total Cost"] = sales["Cant"] * sales["Cost"]
sales["Benefits"] = sales["Ingres"] - (sales["Cost"] * sales["Cant"])
print(sales)
dataSales = pd.DataFrame.from_dict(sales)
print(dataSales)
archiveSales = archiveSales.append(dataSales, ignore_index=True)
print(archiveSales)
#archiveSales.to_excel("Vebook("Ventas.xlsx")
#workSheet1 = workBook.create_sheet("Ventas")
#cambiar el nontas.xlsx")

archiveStock = "Base de datos de stock.xlsx"
archive = pd.read_excel(archiveStock)

#archiveForSales = "Ventas.xlsx"
#archiveSales = pd.read_excel(archiveForSales)

archiveForBuys = "Compras.xlsx"
archiveBuy = pd.read(archiveForBuys)
print(archiveBuy)

#archiveForResume = "Resumen.xlsx"
#archiveResume = pd.read_excel(archiveForResume)
#print(archiveResume)

#resume = {"Date": "", "Tipe": "", "Ingres": "", "Cost": "", "Benefit": ""}
#dateToday = datetime.now()
#dateToday = dateToday.strftime("%d/%m/%Y")
#resume["Date"]= dateToday
#resume["Tipe"]= "Ventas"
#resume["Ingres"] = sales["Ingres"]
#resume["Cost"] = sales["Total Cost"]
#resume["Benefit"] = resume["Ingres"] - resume["Cost"]
#dataResume = pd.DataFrame(resume)
#print(dataResume)